from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# 파일 경로 정의
ny_test_file_path = 'C:/Users/charlton/Desktop/popo.xlsx'

# 파일 이름 분리 및 새 파일 이름 생성
file_name = ny_test_file_path.split('/')[-1]
new_file_name = file_name.replace('.xlsx', ' r1.xlsx')
new_file_path = ny_test_file_path.replace(file_name, new_file_name)

# 엑셀 파일 불러오기
wb = load_workbook(ny_test_file_path)

# 기존 시트의 이름 확인 (첫 번째 시트를 사용한다고 가정)
sheet_name = wb.sheetnames[0]

# 기존 시트를 복사하여 'check'라는 이름의 새로운 시트 생성 (서식 유지)
source = wb[sheet_name]
check = wb.copy_worksheet(source)
check.title = 'check'

# 기존 시트를 복사하여 'pNq_check'라는 이름의 또 다른 새로운 시트 생성 (서식 유지)
pNq_check = wb.copy_worksheet(source)
pNq_check.title = 'pNq_check'

# pandas를 사용하여 데이터 읽기
df = pd.read_excel(ny_test_file_path, sheet_name=sheet_name)

# pandas DataFrame에서 데이터를 'pNq_check' 시트에 추가
for r in dataframe_to_rows(df, index=False, header=True):
    pNq_check.append(r)

# 수정된 워크북을 새 파일로 저장
wb.save(new_file_path)

print(f"새 파일이 저장되었습니다: {new_file_path}")
