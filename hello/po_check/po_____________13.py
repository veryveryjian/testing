import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 파일 경로 정의
ny_test_file_path1 = 'C:/Users/charlton/Desktop/popo.xlsx'
ny_test_file_path2 = 'C:/Users/charlton/Desktop/p.xlsx'

# 첫 번째 파일에서 데이터 읽기 (상위 15행 스킵)
df1 = pd.read_excel(ny_test_file_path1, skiprows=15)

# 두 번째 파일에서 데이터 읽기
df2 = pd.read_excel(ny_test_file_path2)

# df1의 ITEM 열을 기준으로 df2의 ITEM 열과 left join 수행
result_df = pd.merge(df1, df2[['ITEM', 'Price']], on='ITEM', how='left', suffixes=('_x', '_y'))

# Price_y - Price_x 계산
result_df['y-x'] = result_df['Price_y'] - result_df['Price_x']

# Price_y 열과 df1의 두 번째 열(Quantity로 가정)을 곱한 값 계산
# 여기서는 df1의 두 번째 열이 실제로 'Quantity'라고 가정합니다.
result_df['y*qty'] = result_df['Price_y'] * result_df.iloc[:, 1]

# y*qty - Amount 계산
result_df['y*qty - Amount'] = result_df['y*qty'] - result_df['Amount']

# 결과 데이터프레임 출력
print(result_df)

# 엑셀 파일 저장 준비
# 파일 이름 분리 및 새 파일 이름 생성
file_name = ny_test_file_path1.split('/')[-1]
new_file_name = file_name.replace('.xlsx', '_c.xlsx')
new_file_path = ny_test_file_path1.replace(file_name, new_file_name)

# openpyxl을 사용하여 워크북 생성 및 시트에 데이터 추가
wb = load_workbook(ny_test_file_path1)
if 'PnQ_check' in wb.sheetnames:
    ws = wb['PnQ_check']
else:
    ws = wb.create_sheet(title='PnQ_check')

# 결과 데이터프레임의 열 제목 포함하여 시트에 추가
for r in dataframe_to_rows(result_df, index=False, header=True):
    ws.append(r)

# 새 파일로 저장
wb.save(new_file_path)

print(f"결과 파일이 저장되었습니다: {new_file_path}")
