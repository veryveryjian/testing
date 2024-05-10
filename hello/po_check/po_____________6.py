import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows  # 이 부분을 추가하세요.

# 파일 경로 정의
ny_test_file_path = 'C:/Users/charlton/Desktop/popo.xlsx'

# pandas를 사용하여 데이터 읽기
df = pd.read_excel(ny_test_file_path)
print(df)

# openpyxl을 사용하여 파일 불러오기 및 수정
wb = load_workbook(ny_test_file_path)
sheet_name = wb.sheetnames[0]  # 기존 시트 이름
if 'pNq_check' in wb.sheetnames:  # 'pNq_check' 시트가 이미 존재한다면 그 시트를 사용
    ws = wb['pNq_check']
else:
    ws = wb.create_sheet(title='pNq_check')  # 새 시트 생성

# pandas DataFrame에서 데이터를 새 시트에 추가
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# 파일 저장
new_file_name = ny_test_file_path.replace('.xlsx', ' r1.xlsx')
wb.save(new_file_name)
print(f"새 파일이 저장되었습니다: {new_file_name}")
