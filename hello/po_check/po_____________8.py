from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# 파일 경로 정의
ny_test_file_path = 'C:/Users/charlton/Desktop/popo.xlsx'

# 엑셀 파일 불러오기
wb = load_workbook(ny_test_file_path)

# 기존 시트의 이름 확인 (첫 번째 시트를 사용한다고 가정)
sheet_name = wb.sheetnames[0]

# 'pNq_check' 시트가 이미 존재하는지 확인하고, 없으면 생성
if 'pNq_check' in wb.sheetnames:
    pNq_check = wb['pNq_check']  # 존재하면 해당 시트 사용
else:
    pNq_check = wb.create_sheet(title='pNq_check')  # 없으면 새로 생성

# pandas를 사용하여 15행을 제외한 데이터 읽기
df = pd.read_excel(ny_test_file_path, sheet_name=sheet_name, skiprows=15)

# 읽어온 데이터 출력
print(df)

# 읽어온 pandas DataFrame에서 데이터를 'pNq_check' 시트에 추가
for r in dataframe_to_rows(df, index=False, header=True):
    pNq_check.append(r)

# 새 파일 이름 생성 및 저장
file_name = ny_test_file_path.split('/')[-1]
new_file_name = file_name.replace('.xlsx', ' r1.xlsx')
new_file_path = ny_test_file_path.replace(file_name, new_file_name)
wb.save(new_file_path)

print(f"새 파일이 저장되었습니다: {new_file_path}")
