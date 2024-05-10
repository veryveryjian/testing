from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# 파일 경로 정의
ny_test_file_path = 'C:/Users/charlton/Desktop/popo.xlsx'

# 엑셀 파일 불러오기
wb = load_workbook(ny_test_file_path)

# 기존 시트의 이름 확인 (첫 번째 시트를 사용한다고 가정)
sheet_name = wb.sheetnames[0]

# 'check' 시트가 이미 존재하는지 확인하고, 없으면 복사해서 생성
if 'check' not in wb.sheetnames:
    source = wb[sheet_name]
    check = wb.copy_worksheet(source)
    check.title = 'check'

# 'pNq_check' 시트가 이미 존재하는지 확인하고, 없으면 복사해서 생성
if 'pNq_check' not in wb.sheetnames:
    source = wb[sheet_name]
    pNq_check = wb.copy_worksheet(source)
    pNq_check.title = 'pNq_check'

# pandas를 사용하여 15행을 제외한 데이터 읽기
df = pd.read_excel(ny_test_file_path, sheet_name=sheet_name, skiprows=15)

# 'pNq_check' 시트에 DataFrame 데이터 추가
# 이때부터는 openpyxl의 기능을 사용하여 데이터를 하나씩 추가해야 합니다.
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 16):
    for c_idx, value in enumerate(row, 1):
        pNq_check.cell(row=r_idx, column=c_idx, value=value)

# 새 파일 이름 생성 및 저장
file_name = ny_test_file_path.split('/')[-1]
new_file_name = file_name.replace('.xlsx', ' r1.xlsx')
new_file_path = ny_test_file_path.replace(file_name, new_file_name)
wb.save(new_file_path)

print(f"새 파일이 저장되었습니다: {new_file_path}")
