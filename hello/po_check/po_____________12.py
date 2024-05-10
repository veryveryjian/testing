import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 파일 경로 정의
ny_test_file_path1 = 'C:/Users/charlton/Desktop/popo.xlsx'
ny_test_file_path2 = 'C:/Users/charlton/Desktop/p.xlsx'

# 첫 번째 파일에서 데이터 읽기 (상위 15행 스킵)
df1 = pd.read_excel(ny_test_file_path1, skiprows=15)

# 두 번째 열이 정수인지 검증하는 로직
if not all(df1.iloc[:, 1].dropna().apply(lambda x: x.is_integer())):
    raise ValueError("소수점 에러: df1의 두 번째 열에 소수점이 있는 숫자가 있습니다.")

# 두 번째 파일에서 데이터 읽기
df2 = pd.read_excel(ny_test_file_path2)

# df1의 ITEM 열을 기준으로 df2의 ITEM 열과 left join 수행
result_df = pd.merge(df1, df2[['ITEM', 'Price']], on='ITEM', how='left')

# 엑셀 파일 저장 준비
# 파일 이름 분리 및 새 파일 이름 생성
file_name = ny_test_file_path1.split('/')[-1]
new_file_name = file_name.replace('.xlsx', '_c.xlsx')
new_file_path = ny_test_file_path1.replace(file_name, new_file_name)

# openpyxl을 사용하여 워크북 생성 및 시트에 데이터 추가
wb = load_workbook(ny_test_file_path1)
if 'PnQ_check' in wb.sheetnames:
    ws = wb['PnQ_check']
else:
    ws = wb.create_sheet(title='PnQ_check')

# 결과 데이터프레임의 열 제목 포함하여 시트에 추가
for r in dataframe_to_rows(result_df, index=False, header=True):
    ws.append(r)

# 새 파일로 저장
wb.save(new_file_path)

print(f"결과 파일이 저장되었습니다: {new_file_path}")
