from openpyxl import load_workbook

# 파일 경로 정의
ny_test_file_path = 'C:/Users/charlton/Desktop/popo.xlsx'

# 파일 이름 분리 및 새 파일 이름 생성
file_name = ny_test_file_path.split('/')[-1]
new_file_name = file_name.replace('.xlsx', ' r1.xlsx')
new_file_path = ny_test_file_path.replace(file_name, new_file_name)

# 엑셀 파일 불러오기
wb = load_workbook(ny_test_file_path)

# 기존 시트의 이름 확인 (첫 번째 시트를 사용한다고 가정)
sheet_name = wb.sheetnames[0]

# 기존 시트를 복사하여 새로운 시트 생성 (서식 유지)
source = wb[sheet_name]
target = wb.copy_worksheet(source)
target.title = 'check'

# 수정된 워크북을 새 파일로 저장
wb.save(new_file_path)

print(f"새 파일이 저장되었습니다: {new_file_path}")
