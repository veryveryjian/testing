import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# Excel 파일 경로 설정 ('test626.xlsx' 파일 사용)
excel_file = r'C:\Users\charlton\Desktop\test626.xlsx'

# 'INVOICE' 시트에서 B16:G 범위의 데이터 읽어오기
df = pd.read_excel(excel_file, sheet_name='INVOICE', skiprows=15, usecols='B:G')

# 'Item'과 'Pcs' 열만 선택하여 새 DataFrame 생성, Pcs를 숫자형으로 변환
COUNT = df[['Item', 'Pcs']].copy()
COUNT['Pcs'] = pd.to_numeric(COUNT['Pcs'], errors='coerce')

# NaN 아닌 행만 필터링
filtered_COUNT = COUNT.dropna(subset=['Item', 'Pcs'])

# Item 별로 Pcs 합계 계산
grouped_COUNT = filtered_COUNT.groupby('Item', as_index=False)['Pcs'].sum()

# 'Total' 행 추가 및 Pcs 열의 전체 합계 계산
total_pcs = grouped_COUNT['Pcs'].sum()
total_row = pd.DataFrame([{'Item': 'Total', 'Pcs': total_pcs}])

# grouped_COUNT에 total_row 추가
final_COUNT = pd.concat([grouped_COUNT, total_row], ignore_index=True)

# 엑셀 파일에 저장
output_file_count = r'C:\Users\charlton\Desktop\COUNT_final_r1.xlsx'
final_COUNT.to_excel(output_file_count, index=False, sheet_name='COUNT')

# 'INVOICE' 시트에서 G7, G12 값 추출
df_invoice = pd.read_excel(excel_file, sheet_name='INVOICE')
value_g7 = df_invoice.iloc[5, 6]  # G7 셀 값
value_g12 = df_invoice.iloc[10, 6]  # G12 셀 값

# openpyxl로 엑셀 파일 수정
wb = load_workbook(output_file_count)
ws = wb.active

# A1,B1 병합 및 'DATE' 문자열 삽입, G7 및 G12 값 삽입
ws.merge_cells('A1:B1')
ws['A1'] = 'DATE'
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:B2')
ws['A2'] = value_g7
ws['A2'].alignment = Alignment(horizontal='center')

ws.merge_cells('C1:C2')
ws['C1'] = value_g12
ws['C1'].alignment = Alignment(horizontal='center')

# 파일 저장
wb.save(output_file_count)
