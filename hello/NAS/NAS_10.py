from openpyxl import load_workbook

# 파일 경로
file_path = "D:\\SynologyDrive\\SynologyDrive\\test.osheet"

# 엑셀 파일 열기
workbook = load_workbook(file_path)

# 시트 선택 (첫 번째 시트를 선택하거나 시트 이름으로 선택)
sheet = workbook.active  # 첫 번째 시트 선택
# sheet = workbook["Sheet1"]  # 시트 이름으로 선택

# 시트의 각 셀의 값을 출력
for row in sheet.iter_rows(values_only=True):
    for cell in row:
        print(cell)

# 파일 사용이 끝나면 닫기
workbook.close()
