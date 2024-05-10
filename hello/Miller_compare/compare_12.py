import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# 파일 이름 변수 지정
inv_filename = "S174"
ir_filename = "IR4699"

# 파일 경로 변수로 변경
df_inv_path = f"C:\\Users\\charlton\\Desktop\\test\\{inv_filename}.CSV"
df_ir_path = f"C:\\Users\\charlton\\Desktop\\test\\{ir_filename}.CSV"

# CSV 파일을 pandas 데이터프레임으로 읽기
df_inv = pd.read_csv(df_inv_path, encoding='ISO-8859-1')
df_ir = pd.read_csv(df_ir_path, encoding='ISO-8859-1')


# 'Debit'와 'Credit'을 사용해 'Amount' 계산
df_inv['Amount'] = df_inv['Debit'] - df_inv['Credit']

# 필요한 열 선택
df_inv_selected = df_inv[['Item', 'Qty', 'Credit']]
df_ir_selected = df_ir[['Item', 'Qty', 'Amount']]

# 피벗 테이블 생성
inv_pivot = df_inv_selected.groupby('Item').agg({'Qty': 'sum', 'Credit': 'sum'}).reset_index()
ir_pivot = df_ir_selected.groupby('Item').agg({'Qty': 'sum', 'Amount': 'sum'}).reset_index()

# 비교 데이터 생성
inv_vl_compare = pd.merge(inv_pivot[['Item', 'Qty']], ir_pivot[['Item', 'Qty']], on='Item', how='left',
                          suffixes=('_INV', '_IR')).fillna(0)
ir_vl_compare = pd.merge(ir_pivot[['Item', 'Qty']], inv_pivot[['Item', 'Qty']], on='Item', how='left',
                         suffixes=('_IR', '_INV')).fillna(0)

# 비교 데이터에 Total_Qty 열 추가
inv_vl_compare['Total_Qty'] = inv_vl_compare['Qty_INV'] + inv_vl_compare['Qty_IR']
ir_vl_compare['Total_Qty'] = ir_vl_compare['Qty_IR'] + ir_vl_compare['Qty_INV']

# 바탕화면 경로 설정
desktop_path = 'C:\\Users\\charlton\\Desktop\\output_excel.xlsx'

# 엑셀 파일로 저장
with pd.ExcelWriter(desktop_path, engine='openpyxl', mode='w') as writer:
    # 요약 데이터 추가
    summary_data = pd.DataFrame({
        'Sheet': ['INV', 'IR'],
        'Item Count': [df_inv_selected['Item'].count(), df_ir_selected['Item'].count()],  # 중복 포함한 전체 Item 수
        'Qty Sum': [df_inv_selected['Qty'].sum(), df_ir_selected['Qty'].sum()],
        'Credit/Amount Sum': [df_inv_selected['Credit'].sum(), df_ir_selected['Amount'].sum()],
        'Deduplicated Item Count': [len(inv_pivot), len(ir_pivot)],  # 중복 제거 후 Item 카운트
        'Pivot Qty Sum': [inv_pivot['Qty'].sum(), ir_pivot['Qty'].sum()]  # 피벗 테이블 기준 Qty 합계
    })
    summary_data.to_excel(writer, sheet_name='Summary', index=False)

    # 원본 및 중복 제거된 데이터 저장
    df_inv_selected.to_excel(writer, sheet_name='INV_rawdata', index=False)
    df_ir_selected.to_excel(writer, sheet_name='IR_rawdata', index=False)
    inv_pivot.to_excel(writer, sheet_name='INV_VL', index=False)
    ir_pivot.to_excel(writer, sheet_name='IR_VL', index=False)

    # 비교 데이터 저장
    inv_vl_compare.to_excel(writer, sheet_name='INV_VL_Compare', index=False)
    ir_vl_compare.to_excel(writer, sheet_name='IR_VL_Compare', index=False)

# 저장된 엑셀 파일을 openpyxl로 열기
wb = load_workbook(desktop_path)

# INV_VL_Compare 시트에 조건부 포맷 적용
sheet = wb['INV_VL_Compare']
for row in sheet.iter_rows(min_row=2, min_col=sheet.max_column, max_col=sheet.max_column, max_row=sheet.max_row):
    for cell in row:
        if cell.value > 0:
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# IR_VL_Compare 시트에 조건부 포맷 적용
sheet = wb['IR_VL_Compare']
for row in sheet.iter_rows(min_row=2, min_col=sheet.max_column, max_col=sheet.max_column, max_row=sheet.max_row):
    for cell in row:
        if cell.value > 0:
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# 변경사항 저장
wb.save(desktop_path)
