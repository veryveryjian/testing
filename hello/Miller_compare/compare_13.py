import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# 파일 이름 동적 지정을 위한 변수 설정
inv_file_name = "S174"
ir_file_name = "IR4699"

# 파일 경로 동적 생성
df_inv_path = f"C:\\Users\\charlton\\Desktop\\test\\{inv_file_name}.CSV"
df_ir_path = f"C:\\Users\\charlton\\Desktop\\test\\{ir_file_name}.CSV"

# CSV 파일을 pandas 데이터프레임으로 읽기
df_inv = pd.read_csv(df_inv_path, encoding='ISO-8859-1')
df_ir = pd.read_csv(df_ir_path, encoding='ISO-8859-1')

# 필요한 열 선택
df_inv_selected = df_inv[['Item', 'Qty', 'Credit']]
df_ir_selected = df_ir[['Item', 'Qty', 'Amount']]

# 피벗 테이블 생성
inv_pivot = df_inv_selected.groupby('Item').agg({'Qty': 'sum', 'Credit': 'sum'}).reset_index()
ir_pivot = df_ir_selected.groupby('Item').agg({'Qty': 'sum', 'Amount': 'sum'}).reset_index()

# 비교 데이터 생성
inv_vl_compare = pd.merge(inv_pivot[['Item', 'Qty']], ir_pivot[['Item', 'Qty']], on='Item', how='left', suffixes=('_INV', '_IR')).fillna(0)
ir_vl_compare = pd.merge(ir_pivot[['Item', 'Qty']], inv_pivot[['Item', 'Qty']], on='Item', how='left', suffixes=('_IR', '_INV')).fillna(0)

# 비교 데이터에 Total_Qty 열 추가
inv_vl_compare['Total_Qty'] = inv_vl_compare['Qty_INV'] + inv_vl_compare['Qty_IR']
ir_vl_compare['Total_Qty'] = ir_vl_compare['Qty_IR'] + ir_vl_compare['Qty_INV']

# 동적으로 생성된 파일 이름
output_file_name = f"{inv_file_name}_{ir_file_name}.xlsx"
desktop_path = f'C:\\Users\\charlton\\Desktop\\{output_file_name}'

# 엑셀 파일로 저장
with pd.ExcelWriter(desktop_path, engine='openpyxl', mode='w') as writer:
    df_inv_selected.to_excel(writer, sheet_name='INV_rawdata', index=False)
    df_ir_selected.to_excel(writer, sheet_name='IR_rawdata', index=False)
    inv_pivot.to_excel(writer, sheet_name='INV_VL', index=False)
    ir_pivot.to_excel(writer, sheet_name='IR_VL', index=False)
    inv_vl_compare.to_excel(writer, sheet_name='INV_VL_Compare', index=False)
    ir_vl_compare.to_excel(writer, sheet_name='IR_VL_Compare', index=False)

# 저장된 엑셀 파일을 openpyxl로 열기 및 열 너비 설정
wb = load_workbook(desktop_path)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for col in range(1, 11):  # A-J 열에 해당
        ws.column_dimensions[get_column_letter(col)].width = 15
wb.save(desktop_path)
