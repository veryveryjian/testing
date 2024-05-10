import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

# Excel 파일 경로 설정
excel_file = r'C:\Users\charlton\Desktop\CI\PO#330MS  INVOICE（TCKU6328636）.xlsx'

# 'INVOICE' 시트에서 B16:G 범위의 데이터 읽어오기
df = pd.read_excel(excel_file, sheet_name='INVOICE', skiprows=15, usecols='B:G')

# 'Item'과 'Pcs' 열만 선택하여 새 DataFrame 생성, Pcs를 숫자형으로 변환
COUNT = df[['Item', 'Pcs']].copy()
COUNT['Pcs'] = pd.to_numeric(COUNT['Pcs'], errors='coerce')

# NaN 아닌 행만 필터링
filtered_COUNT = COUNT.dropna(subset=['Item', 'Pcs'])

# Item 별로 Pcs 합계 계산
grouped_COUNT = filtered_COUNT.groupby('Item', as_index=False)['Pcs'].sum()

# 'Total' 행 추가 및 Pcs 열의 전체 합계 계산
total_pcs = grouped_COUNT['Pcs'].sum()
total_row = pd.DataFrame([{'Item': 'Total', 'Pcs': total_pcs}])

# grouped_COUNT에 total_row 추가
final_COUNT = pd.concat([grouped_COUNT, total_row], ignore_index=True)

# 엑셀 파일에 저장
output_file_count = r'C:\Users\charlton\Desktop\CI\COUNT_final_r3.xlsx'
final_COUNT.to_excel(output_file_count, index=False, sheet_name='COUNT')

# 'INVOICE' 시트에서 G7, G12 값 추출
df_invoice = pd.read_excel(excel_file, sheet_name='INVOICE')
value_g7 = df_invoice.iloc[5, 6]  # G7 셀 값
value_g12 = df_invoice.iloc[10, 6]  # G12 셀 값

# openpyxl로 엑셀 파일 수정
wb = load_workbook(output_file_count)
ws = wb.active

# A, B, C 컬럼 너비 설정
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 25

# 테두리 스타일 정의
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# 데이터가 있는 셀에 테두리 적용
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border

# 병합 및 값 삽입
ws.merge_cells('A1:B1')
ws['A1'] = 'DATE'
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:B2')
ws['A2'] = value_g7
ws['A2'].alignment = Alignment(horizontal='center')

ws.merge_cells('C1:C2')
ws['C1'] = value_g12
ws['C1'].alignment = Alignment(horizontal='center')

# 파일 저장
wb.save(output_file_count)
