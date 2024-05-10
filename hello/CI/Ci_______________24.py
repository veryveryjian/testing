from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
from openpyxl.styles import Alignment


# 엑셀 파일 로드
excel_file = r'C:\Users\charlton\Desktop\CI\PO#330MS  INVOICE（TCKU6328636）.xlsx'
wb = load_workbook(excel_file)
ws = wb.active  # 활성 시트

# 새 시트 생성 또는 기존 시트 초기화
new_sheet_names = ['RECEIPT', 'COUNT', 'MATCH']
for sheet_name in new_sheet_names:
    if sheet_name in wb:
        del wb[sheet_name]
    wb.create_sheet(sheet_name)

# 셀 복사, 병합, 셀 너비, 텍스트 래핑 해제 반영
for new_sheet_name in new_sheet_names:
    new_ws = wb[new_sheet_name]

    # 셀 데이터 및 서식 복사
    for row in ws.iter_rows():
        for cell in row:
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

                # 0번째 행부터 16번째 행까지 텍스트 래핑 해제
                if cell.row <= 17:
                    new_cell.alignment = Alignment(wrap_text=False)

    # 셀 병합 상태 복사
    for merge_cell in ws.merged_cells:
        new_ws.merge_cells(str(merge_cell))

    # 셀 너비 복사
    for col in ws.columns:
        new_ws.column_dimensions[get_column_letter(col[0].column)].width = ws.column_dimensions[
            get_column_letter(col[0].column)].width

# 변경 사항 저장
wb.save(r'C:\Users\charlton\Desktop\CI\PO#330MS  INVOICE（TCKU6328636）_modified_r4.xlsx')
