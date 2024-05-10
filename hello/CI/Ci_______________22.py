from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# 엑셀 파일 로드
excel_file = r'C:\Users\charlton\Desktop\CI\PO#330MS  INVOICE（TCKU6328636）.xlsx'
wb = load_workbook(excel_file)
ws = wb.active  # 활성 시트

# 새 시트 이름들
new_sheets = ['RECEIPT', 'COUNT', 'MATCH']

for new_sheet_name in new_sheets:
    # 이미 존재하는 시트가 있으면 삭제
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]
    new_ws = wb.create_sheet(new_sheet_name)  # 새 시트 생성

    # 셀 데이터와 서식 복사
    for row in ws.iter_rows():
        for cell in row:
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)

            # 셀 서식 복사
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

# 변경 사항 저장
wb.save(r'C:\Users\charlton\Desktop\CI\PO#330MS  INVOICE（TCKU6328636）_modified_R2.xlsx')
