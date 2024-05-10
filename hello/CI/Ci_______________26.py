import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from copy import copy

# Excel 파일 경로 설정
excel_file = r'C:\Users\charlton\Desktop\CI\PO#330MS  INVOICE（TCKU6328636）.xlsx'

# 데이터 처리 부분 (COUNT DataFrame 생성 및 처리)
df = pd.read_excel(excel_file, sheet_name='INVOICE', skiprows=15, usecols='B:G')
COUNT = df[['Item', 'Pcs']].copy()
COUNT['Pcs'] = pd.to_numeric(COUNT['Pcs'], errors='coerce')
filtered_COUNT = COUNT.dropna(subset=['Item', 'Pcs'])
grouped_COUNT = filtered_COUNT.groupby('Item', as_index=False)['Pcs'].sum()
total_pcs = grouped_COUNT['Pcs'].sum()
total_row = pd.DataFrame([{'Item': 'Total', 'Pcs': total_pcs}])
final_COUNT = pd.concat([grouped_COUNT, total_row], ignore_index=True)

# 엑셀 파일 로드 및 수정
wb = load_workbook(excel_file)

# 'COUNT' 시트 처리
if 'COUNT' in wb.sheetnames:
    del wb['COUNT']
ws_count = wb.create_sheet('COUNT')
for r_idx, row in enumerate(pd.DataFrame(final_COUNT).itertuples(index=False), 1):
    for c_idx, value in enumerate(row, 1):
        ws_count.cell(row=r_idx, column=c_idx).value = value

# 새 시트 이름들 처리 ('RECEIPT', 'MATCH', 'PRINT')
new_sheets = ['RECEIPT', 'MATCH', 'PRINT']
for new_sheet_name in new_sheets:
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]
    new_ws = wb.create_sheet(new_sheet_name)

    # 활성 시트의 데이터 및 서식 복사
    for row in wb.active.iter_rows():
        for cell in row:
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # 병합된 셀 및 셀 너비 복사
    for merge_range in wb.active.merged_cells.ranges:
        new_ws.merge_cells(str(merge_range))
    for col in wb.active.columns:
        col_letter = get_column_letter(col[0].column)
        new_ws.column_dimensions[col_letter].width = wb.active.column_dimensions[col_letter].width

# 변경 사항 저장
wb.save(r'C:\Users\charlton\Desktop\CI\PO#330MS  INVOICE（TCKU6328636）_modified_r10.xlsx')
