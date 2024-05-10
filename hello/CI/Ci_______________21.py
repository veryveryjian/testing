from openpyxl import load_workbook

# 엑셀 파일 경로
excel_file = r'C:\Users\charlton\Desktop\CI\PO#330MS  INVOICE（TCKU6328636）.xlsx'

# 엑셀 파일 로드
wb = load_workbook(excel_file)
ws = wb.active  # 활성 시트

# 'L' 열 뒤의 모든 열 제거
max_col = ws.max_column  # 시트의 최대 열 수
col_L_index = ws['L'][0].column  # 'L' 열의 인덱스 (1부터 시작)

# 'L' 열 뒤의 열들을 제거
if col_L_index < max_col:
    ws.delete_cols(col_L_index + 1, max_col - col_L_index)

# 적용된 내용을 새 시트(RECEIPT, COUNT, MATCH)에 복사
for sheet_name in ['RECEIPT', 'COUNT', 'MATCH']:
    if sheet_name in wb.sheetnames:  # 이미 시트가 존재하면 삭제
        wb.remove(wb[sheet_name])
    wb.create_sheet(sheet_name)  # 새 시트 생성
    new_ws = wb[sheet_name]

    # 데이터 복사 및 붙여넣기
    for row in ws.iter_rows(values_only=True):
        new_ws.append(row)

# 변경 사항 저장
wb.save(r'C:\Users\charlton\Desktop\CI\PO#330MS  INVOICE（TCKU6328636）_modified_R1.xlsx')
