from openpyxl import load_workbook
from copy import copy

# 엑셀 파일 로드
excel_file = r'C:\Users\charlton\Desktop\CI\PO#330MS  INVOICE（TCKU6328636）.xlsx'
wb = load_workbook(excel_file)
ws = wb.active  # 활성 시트

# 새 시트 생성 또는 기존 시트 초기화
new_sheets = ['RECEIPT', 'COUNT', 'MATCH']
for sheet_name in new_sheets:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    wb.create_sheet(sheet_name)

for new_sheet_name in new_sheets:
    new_ws = wb[new_sheet_name]

    # 셀 데이터와 서식 복사
    for row in ws.iter_rows():
        for cell in row:
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # 병합된 셀 복사
    for range_ in ws.merged_cells.ranges:
        new_ws.merge_cells(str(range_))

    # wrapText 속성이 적용된 셀 복사 (이미 위에서 alignment를 포함해 복사했으므로 별도 처리 필요 없음)

# 변경 사항 저장
wb.save(r'C:\Users\charlton\Desktop\CI\PO#330MS  INVOICE（TCKU6328636）_modified_R3.xlsx')
