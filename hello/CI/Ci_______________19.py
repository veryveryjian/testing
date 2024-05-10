from openpyxl import load_workbook

# 엑셀 파일 경로
excel_file = r'C:\Users\charlton\Desktop\CI\PO#330MS  INVOICE（TCKU6328636）.xlsx'

# 엑셀 파일 로드
wb = load_workbook(excel_file)
ws = wb.active  # 활성 시트

# 'L' 열 뒤의 모든 열 제거
max_col = ws.max_column  # 시트의 최대 열 수
col_L_index = ws['L'][0].column  # 'L' 열의 인덱스 (1부터 시작)

# 'L' 열 뒤의 열들을 제거하기 위해 반복문 사용
for col in range(max_col, col_L_index, -1):  # max_col부터 L열까지 역순으로
    ws.delete_cols(col)

# 변경 사항 저장
wb.save(r'C:\Users\charlton\Desktop\CI\PO#330MS  INVOICE（TCKU6328636）_modified.xlsx')





