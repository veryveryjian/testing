from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from copy import copy

# Excel 파일 로드
excel_file = r'C:\Users\charlton\Desktop\CI\PO#330MS  INVOICE（TCKU6328636）.xlsx'
wb = load_workbook(excel_file)
ws = wb.active  # 활성 시트

# 새 시트 이름들
new_sheets = ['RECEIPT', 'MATCH','PRINT']

for new_sheet_name in new_sheets:
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]  # 기존 시트 삭제
    new_ws = wb.create_sheet(new_sheet_name)  # 새 시트 생성

    # 셀 데이터 및 서식 복사
    for row in ws.iter_rows():
        for cell in row:
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # 병합된 셀 복사 (수정된 부분)
    for merge_range in ws.merged_cells.ranges:
        new_ws.merge_cells(str(merge_range))

    # 셀 너비 복사
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        new_ws.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width

    # G10 셀에만 text wrap 적용
    new_ws['G10'].alignment = Alignment(wrap_text=True)

    # 0-5 행에 middle align과 center 정렬 적용
    for row in new_ws.iter_rows(min_row=1, max_row=6):
        for cell in row:
            cell.alignment = Alignment(vertical='center', horizontal='center')

# 변경 사항 저장
wb.save(r'C:\Users\charlton\Desktop\CI\PO#330MS  INVOICE（TCKU6328636）_modified_r6.xlsx')
